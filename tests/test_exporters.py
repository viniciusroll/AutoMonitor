"""Testes dos exportadores (CSV, Excel, JSON)."""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.exceptions import ExportError
from app.exporters import get_exporter
from app.exporters.base import EXPORT_COLUMNS
from app.services.vehicle_service import VehicleService


@pytest.fixture()
def vehicles(session, scraped_factory):
    service = VehicleService(session)
    service.save_many(
        [
            scraped_factory(external_id="1", title="Honda Civic", price=95_000),
            scraped_factory(external_id="2", title="Toyota Corolla", price=110_000),
        ]
    )
    return list(service.list_vehicles())


def test_export_csv(tmp_path: Path, vehicles) -> None:
    path = tmp_path / "out.csv"
    get_exporter("csv").export(vehicles, path=path)
    with path.open(encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 2
    assert list(rows[0].keys()) == list(EXPORT_COLUMNS)


def test_export_json(tmp_path: Path, vehicles) -> None:
    path = tmp_path / "out.json"
    get_exporter("json").export(vehicles, path=path)
    data = json.loads(path.read_text(encoding="utf-8"))
    assert len(data) == 2
    assert {v["title"] for v in data} == {"Honda Civic", "Toyota Corolla"}


def test_export_excel(tmp_path: Path, vehicles) -> None:
    path = tmp_path / "out.xlsx"
    get_exporter("excel").export(vehicles, path=path)
    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.max_row == 3  # cabeçalho + 2 linhas
    assert [c.value for c in sheet[1]] == list(EXPORT_COLUMNS)


def test_formato_invalido() -> None:
    with pytest.raises(ExportError):
        get_exporter("pdf")
