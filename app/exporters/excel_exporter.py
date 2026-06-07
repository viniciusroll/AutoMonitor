"""Exportador para Excel (.xlsx) usando OpenPyXL."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import ExportError
from app.exporters.base import EXPORT_COLUMNS, BaseExporter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_MAX_WIDTH = 60


class ExcelExporter(BaseExporter):
    """Gera uma planilha .xlsx formatada com cabeçalho destacado."""

    extension = "xlsx"

    def _write(self, rows: Sequence[dict[str, object]], path: Path) -> None:
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Veículos"
            self._write_header(sheet)
            for row in rows:
                sheet.append([row.get(col) for col in EXPORT_COLUMNS])
            self._autosize(sheet, rows)
            sheet.freeze_panes = "A2"
            workbook.save(path)
        except OSError as exc:
            raise ExportError(f"Falha ao escrever Excel em {path}: {exc}") from exc

    @staticmethod
    def _write_header(sheet: Worksheet) -> None:
        sheet.append(list(EXPORT_COLUMNS))
        for cell in sheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

    @staticmethod
    def _autosize(sheet: Worksheet, rows: Sequence[dict[str, object]]) -> None:
        for index, column in enumerate(EXPORT_COLUMNS, start=1):
            longest = len(column)
            for row in rows:
                value = row.get(column)
                if value is not None:
                    longest = max(longest, len(str(value)))
            sheet.column_dimensions[get_column_letter(index)].width = min(
                longest + 2, _MAX_WIDTH
            )
